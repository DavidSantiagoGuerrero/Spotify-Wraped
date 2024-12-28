import openpyxl as xl

def BorrarDuplicados(archivo, tabla, filas, numFilas):
    i = 0
    while i < numFilas - 1:  # Usamos un bucle while para evitar saltos inesperados
        if filas[i] == filas[i + 1]:
            tabla.delete_rows(i + 2)  # i+2 porque las filas en Excel son 1-indexadas
            numFilas -= 1
            filas.pop(i + 1)  # Eliminamos el duplicado de la lista también
        else:
            i += 1  # Avanzamos solo si no eliminamos filas
    archivo.save('Spotify Wrapped.xlsx')  # Guardamos los cambios en el archivo

try:
	excel = xl.load_workbook('Spotify Wrapped.xlsx')
except FileNotFoundError:
	print("The file was not found. Please check the file path.")
	exit()

xd = excel.active

UltimaColumna = xd.max_column
UltimaFila = xd.max_row

print('Total number of rows: '+str(xd.max_row)+'. And total number of columns: '+str(xd.max_column))

values = [xd.cell(row=1,column=i+1).value for i in range(0, UltimaColumna)]

canciones = [xd.cell(row=i+1,column=2).value for i in range(1, UltimaFila)]

print(values)
print(canciones)
BorrarDuplicados(excel, xd, canciones, UltimaFila-1)